from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import calendar

MONTHS = ["Sausis", "Vasaris", "Kovas", "Balandis", "Gegužė", "Birželis", "Liepa", "Rugpjūtis", "Rugsėjis", "Spalis", "Lapkritis", "Gruodis",]

def input_year():
    print(" Įveskite datą:\n")
    while True:
        y = input("   Metus:").strip()
        if y.isdigit() and len(y) == 4:
            return int(y)
        print("Netinkami metai! Įveskite keturženklį skaičių, pvz., 2026.")

def input_month():
    while True:
        m = input("   Mėnesį: ").strip().lower()
        if m.isdigit():
            i = int(m)
            if 1 <= i <= 12:
                return MONTHS[i - 1]
        if len(m) >= 4:
            for month in MONTHS:
                if month.lower().startswith(m[:4]):
                    return month
        print("Netinkamas mėnuo. Pvz: 8, 08, rugpjutis, Rugpjūtis")

def input_day(year, month_name):
    month_num = None
    for i, name in enumerate(MONTHS, start=1):
        if name.lower() == month_name.lower():
            month_num = i
            break
    if month_num is None:
        month_num = 1
    _, max_day = calendar.monthrange(year, month_num)
    while True:
        d = input(f"Diena (1-{max_day}): ").strip()
        if d.isdigit():
            day = int(d)
            if 1 <= day <= max_day:
                return day
        print(f"Netinkama diena! Įveskite skaičių nuo 1 iki {max_day}.")

def input_hours(prompt):
    while True:
        h = input(f"{prompt}: ").strip().replace(',', '.')
        try:
            val = float(h)
            if val >= 0:
                return val
            else:
                print("Negalima įvesti neigiamų valandų!")
        except ValueError:
            print("Netinkamas formatas! Įveskite skaičių, pvz.: 3.5")

def load_objects(filename="objektai.txt"):
    if not os.path.exists(filename):
        return []
    with open(filename, "r", encoding="utf-8") as f:
        # Strip empty lines and trailing spaces
        return [line.strip() for line in f if line.strip()]
    
def input_object(objects):
    min_chars = 5
    while True:
        inp = input(f"Objekto pavadinimas (pirmi {min_chars} simboliai): ").strip()
        if len(inp) < min_chars:
            print(f"Įveskite bent {min_chars} simbolius")
            continue
        
        matches = [obj for obj in objects if obj.lower().startswith(inp.lower())]
        if len(matches) == 1:
            return matches[0]
        elif len(matches) > 1:
            print("Rasta keli objektai: " + ", ".join(matches[:5]) + (", ..." if len(matches) > 5 else ""))
            min_chars = len(inp) + 1
            print(f"Įveskite daugiau simbolių, kad būtų unikalus: (dabar reikia bent {min_chars})")
        else:
            print("Objektas nerastas, bandykite dar kartą.")

def setup_console():
    os.system("title Darbo ataskaitos įrankis")
    os.system("color 8F")

def setup():
    year = input_year()
    month = input_month()
    name = input("   Serviso inžinieriaus v. pavardė (pvz.: Antano Antanausko): ")
    return year, month, name

def unique_object_worktypes(records):
    seen = set()
    result = []
    for r in records:
        key = (r["Darbo tipas"], r["Objektas"])
        if key not in seen:
            seen.add(key)
            result.append(key)
    return result

def append_summary_table(ws, records):
    ws.append([])
    ws.append([])
    ws.append([])

    ws.append([
        "Darbo tipas",
        "Objektas",
        "Pradirbtos valandos",
        "Kelionės laikas",
        "AX S.C.",
        "Viršvalandžiai"
    ])

    start_row = ws.max_row
    combos = unique_object_worktypes(records)

    for work_type, obj in combos:
        row = ws.max_row + 1

        ws.append([
            work_type,
            obj,
            f'=SUMIFS(Ataskaita[Pradirbtos valandos], Ataskaita[Objektas], B{row}, Ataskaita[Darbo tipas (S - šefmontažas, G - garantinis remontas, SP - serviso projektas, V - vizitas, 0 - nei vienas)], A{row})',
            f'=SUMIFS(Ataskaita[Kelionės laikas], Ataskaita[Objektas], B{row}, Ataskaita[Darbo tipas (S - šefmontažas, G - garantinis remontas, SP - serviso projektas, V - vizitas, 0 - nei vienas)], A{row})',
            "",
            f'=SUMIFS(Ataskaita[Viršvalandžiai], Ataskaita[Objektas], B{row}, Ataskaita[Darbo tipas (S - šefmontažas, G - garantinis remontas, SP - serviso projektas, V - vizitas, 0 - nei vienas)], A{row})',
        ])

    end_row = ws.max_row

    table = Table(displayName="Summary", ref=f"A{start_row}:F{end_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ("C", "D", "F"):
        for cell in ws[f"{col}{start_row+1}:{col}{end_row}"]:
            cell[0].number_format = "0.0"
            cell[0].alignment = Alignment(horizontal="left")

def create_excel(year, month, name, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ataskaita"

    ws["A2"] = "Data:"
    ws["B2"] = f"{year}, {month}"

    ws["A2"].font = Font(bold=True)

    ws.merge_cells("A4:E4")
    ws["A4"] = f"Serviso inžinieriaus {name} neaktuotų valandų ataskaita"
    ws["A4"].font = Font(size=14, bold=True)
    ws["A4"].alignment = Alignment(horizontal="center")

    headers = [
        "Diena",
        "Objektas",
        "Pradirbtos valandos",
        "Kelionės laikas",
        "Viršvalandžiai",
        "Darbo tipas (S - šefmontažas, G - garantinis remontas, SP - serviso projektas, V - vizitas, 0 - nei vienas)"
    ]

    ws.append([])
    ws.append(headers)

    for row in records:
        ws.append([
            row["Diena"],
            row["Objektas"],
            row["Pradirbtos valandos"],
            row["Kelionės laikas"],
            row["Viršvalandžiai"],
            row["Darbo tipas"]
        ])

    start_row = 6
    end_row = start_row + len(records)
    table = Table(displayName="Ataskaita", ref=f"A6:F{end_row}")

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30

    append_summary_table(ws, records)

    filename = f"Neaktuotos_valandos_{year}_{month}.xlsx"
    wb.save(filename)

    print(f"Sukurta Excel ataskaita: {filename}")

def data_list(year, month):
    all_data = []
    while True:
        data = {}
        data["Diena"] = input_day(year, month)
        objects_list = load_objects("objektai.txt")
        data["Objektas"] = input_object(objects_list)
        data["Pradirbtos valandos"] = input_hours("Pradirbtos valandos")
        data["Kelionės laikas"] = input_hours("Kelionės laikas")
        data["Viršvalandžiai"] = input_hours("Viršvalandžiai")
        
        while True:
            work_type = input(
                "S - šefmontažas, G - garantinis remontas, SP - serviso projektas, V - vizitas, 0 - nei vienas: "
            ).strip().upper()
            if work_type in {"S", "G", "SP", "V", "0"}:
                data["Darbo tipas"] = work_type
                break
            print("Netinkamas simbolis..(S, G, SP, V, 0)")
        
        all_data.append(data)
        more = input("Sekantis įrašas? (y/n): ").strip().lower()
        if more == "n":
            break

    return all_data

def main():
    setup_console()
    year, month, name = setup()
    records = data_list(year, month)
    create_excel(year, month, name, records)
    input("Spausti ENTER, kad uždaryti...")

if __name__ == "__main__":
    main()